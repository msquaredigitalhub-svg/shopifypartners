"""
Shopify Partner Directory Scraper
==================================

Scrapes the Shopify Service Partners directory and writes the results
to an Excel file.  Can be used as a CLI script or imported as a module
(used by the Flask web app in app.py).

CLI usage:
    pip install requests beautifulsoup4 openpyxl
    python shopify_partner_scraper.py

Web usage:
    pip install flask requests beautifulsoup4 openpyxl
    python app.py
"""

from __future__ import annotations

import re
import time
import random
import signal
import sys
import math
import uuid
import threading
import tempfile
import os
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

BASE_LIST_URL = "https://www.shopify.com/partners/directory/services"
MAX_PAGES = 302
DELAY_SECONDS = (1.5, 3.0)
OUTPUT_FILE = os.path.join(tempfile.gettempdir(), "shopify_partners.xlsx")
REQUEST_TIMEOUT = 20
MAX_RETRIES = 3
MAX_REQUESTS_PER_MINUTE = 30
PARTNERS_PER_PAGE = 16

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

COUNTRIES = {
    "ad": {"name": "Andorra"},
    "ar": {"name": "Argentina"},
    "au": {"name": "Australia"},
    "at": {"name": "Austria"},
    "bh": {"name": "Bahrain"},
    "bd": {"name": "Bangladesh"},
    "by": {"name": "Belarus"},
    "be": {"name": "Belgium"},
    "ba": {"name": "Bosnia & Herzegovina"},
    "br": {"name": "Brazil"},
    "bg": {"name": "Bulgaria"},
    "ca": {"name": "Canada"},
    "cl": {"name": "Chile"},
    "cn": {"name": "China"},
    "co": {"name": "Colombia"},
    "cy": {"name": "Cyprus"},
    "cz": {"name": "Czech Republic"},
    "dk": {"name": "Denmark"},
    "ec": {"name": "Ecuador"},
    "eg": {"name": "Egypt"},
    "ee": {"name": "Estonia"},
    "fi": {"name": "Finland"},
    "fr": {"name": "France"},
    "de": {"name": "Germany"},
    "gr": {"name": "Greece"},
    "gt": {"name": "Guatemala"},
    "hk": {"name": "Hong Kong SAR"},
    "hu": {"name": "Hungary"},
    "is": {"name": "Iceland"},
    "in": {"name": "India"},
    "id": {"name": "Indonesia"},
    "ie": {"name": "Ireland"},
    "il": {"name": "Israel"},
    "it": {"name": "Italy"},
    "jp": {"name": "Japan"},
    "xk": {"name": "Kosovo"},
    "kw": {"name": "Kuwait"},
    "lv": {"name": "Latvia"},
    "lb": {"name": "Lebanon"},
    "lt": {"name": "Lithuania"},
    "lu": {"name": "Luxembourg"},
    "my": {"name": "Malaysia"},
    "mx": {"name": "Mexico"},
    "ma": {"name": "Morocco"},
    "np": {"name": "Nepal"},
    "nl": {"name": "Netherlands"},
    "nz": {"name": "New Zealand"},
    "ng": {"name": "Nigeria"},
    "no": {"name": "Norway"},
    "pk": {"name": "Pakistan"},
    "pa": {"name": "Panama"},
    "pe": {"name": "Peru"},
    "ph": {"name": "Philippines"},
    "pl": {"name": "Poland"},
    "pt": {"name": "Portugal"},
    "ro": {"name": "Romania"},
    "ru": {"name": "Russia"},
    "sm": {"name": "San Marino"},
    "rs": {"name": "Serbia"},
    "sg": {"name": "Singapore"},
    "sk": {"name": "Slovakia"},
    "za": {"name": "South Africa"},
    "kr": {"name": "South Korea"},
    "es": {"name": "Spain"},
    "lk": {"name": "Sri Lanka"},
    "se": {"name": "Sweden"},
    "ch": {"name": "Switzerland"},
    "tw": {"name": "Taiwan"},
    "th": {"name": "Thailand"},
    "tr": {"name": "Türkiye"},
    "ua": {"name": "Ukraine"},
    "ae": {"name": "United Arab Emirates"},
    "gb": {"name": "United Kingdom"},
    "us": {"name": "United States"},
    "vn": {"name": "Vietnam"},
}

# ---------------------------------------------------------------------------
# FILTER CONSTANTS
# ---------------------------------------------------------------------------

INDUSTRIES = {
    "": {"name": "Any Industry"},
    "art_photography": {"name": "Art & Photography"},
    "automotive": {"name": "Automotive"},
    "business_to_business_b2b": {"name": "Business to Business (B2B)"},
    "clothing_fashion": {"name": "Clothing & Fashion"},
    "consumer_packaged_goods": {"name": "Consumer Packaged Goods"},
    "electronics": {"name": "Electronics"},
    "food_drink": {"name": "Food & Drink"},
    "health_beauty": {"name": "Health & Beauty"},
    "home_garden": {"name": "Home & Garden"},
    "jewelry_accessories": {"name": "Jewelry & Accessories"},
    "lifestyle": {"name": "Lifestyle"},
    "pet_care": {"name": "Pet Care"},
    "services": {"name": "Services"},
    "sports_recreation": {"name": "Sports & Recreation"},
    "toys_games": {"name": "Toys & Games"},
}

PARTNER_TIERS = {
    "": {"name": "Any Tier"},
    "tier_plus": {"name": "Plus Partners"},
    "tier_premier": {"name": "Premier Partners"},
    "tier_platinum": {"name": "Platinum Partners"},
}

LANGUAGES = {
    "": {"name": "Any Language"},
    "lang-en": {"name": "English"},
    "lang-es": {"name": "Spanish"},
    "lang-de": {"name": "German"},
    "lang-fr": {"name": "French"},
    "lang-it": {"name": "Italian"},
    "lang-zh_cn": {"name": "Chinese (PRC)"},
    "lang-zh_hk": {"name": "Chinese (Traditional)"},
    "lang-ja": {"name": "Japanese"},
    "lang-pt_br": {"name": "Portuguese (Brazil)"},
    "lang-pt_pt": {"name": "Portuguese (European)"},
    "lang-ru": {"name": "Russian"},
    "lang-he": {"name": "Hebrew"},
    "lang-hi": {"name": "Hindi"},
    "lang-ur": {"name": "Urdu"},
    "lang-ar": {"name": "Arabic"},
    "lang-nl": {"name": "Dutch"},
    "lang-sv": {"name": "Swedish"},
    "lang-no": {"name": "Norwegian"},
    "lang-bn": {"name": "Bengali"},
    "lang-la": {"name": "Lahnda"},
    "lang-ko": {"name": "Korean"},
    "lang-te": {"name": "Telugu"},
    "lang-mr": {"name": "Marathi"},
    "lang-tr": {"name": "Turkish"},
    "lang-ta": {"name": "Tamil"},
    "lang-fi": {"name": "Finnish"},
    "lang-da": {"name": "Danish"},
    "lang-jv": {"name": "Javanese"},
    "lang-th": {"name": "Thai"},
    "lang-id": {"name": "Bahasa Indonesia"},
    "lang-pl": {"name": "Polish"},
    "lang-gu": {"name": "Gujarati"},
}

# ---------------------------------------------------------------------------
# RATE LIMITER
# ---------------------------------------------------------------------------

class RateLimiter:
    def __init__(self, max_per_minute: int = MAX_REQUESTS_PER_MINUTE):
        self.max_per_minute = max_per_minute
        self._timestamps: list[float] = []
        self._lock = threading.Lock()

    def wait_if_needed(self):
        now = time.time()
        with self._lock:
            self._timestamps = [t for t in self._timestamps if now - t < 60]
            if len(self._timestamps) >= self.max_per_minute:
                sleep_time = self._timestamps[0] + 60 - now
                if sleep_time > 0:
                    time.sleep(sleep_time + random.uniform(0, 0.5))
            self._timestamps.append(time.time())

rate_limiter = RateLimiter()

# ---------------------------------------------------------------------------
# SESSION
# ---------------------------------------------------------------------------

session = requests.Session()
session.headers.update(HEADERS)

_stop_requested = False

def _handle_sigint(signum, frame):
    global _stop_requested
    print("\nStop requested — finishing current operation then saving progress...")
    _stop_requested = True

signal.signal(signal.SIGINT, _handle_sigint)

# ---------------------------------------------------------------------------
# URL HELPERS
# ---------------------------------------------------------------------------

def build_list_url(
    country_codes: list[str] | None = None,
    page: int = 1,
    industry: str = "",
    partner_tiers: list[str] | None = None,
    language: str = "",
    sort: str = "DEFAULT",
) -> str:
    params = []
    if country_codes:
        for cc in country_codes:
            params.append(f"locationCodes=loc-{cc}")
    if industry:
        params.append(f"industryHandles={industry}")
    if partner_tiers:
        for tier in partner_tiers:
            if tier:
                params.append(f"partnerTiers={tier}")
    if language:
        params.append(f"languageCodes={language}")
    params.append(f"page={page}")
    params.append(f"sort={sort}")
    return BASE_LIST_URL + "?" + "&".join(params)

# ---------------------------------------------------------------------------
# HTTP FETCH
# ---------------------------------------------------------------------------

def fetch(url: str) -> str | None:
    for attempt in range(1, MAX_RETRIES + 1):
        rate_limiter.wait_if_needed()
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 200:
                return resp.text
            if resp.status_code == 429:
                retry_after = resp.headers.get("Retry-After")
                if retry_after and retry_after.isdigit():
                    wait = int(retry_after)
                else:
                    wait = 10 * attempt
                print(f"  Rate limited on {url}, waiting {wait}s...")
                time.sleep(wait)
                continue
            print(f"  Got status {resp.status_code} for {url}")
            return None
        except requests.RequestException as e:
            print(f"  Request error ({attempt}/{MAX_RETRIES}) for {url}: {e}")
            time.sleep(3 * attempt)
    return None


def polite_sleep():
    time.sleep(random.uniform(*DELAY_SECONDS))

# ---------------------------------------------------------------------------
# CHECK LISTING COUNT
# ---------------------------------------------------------------------------

def check_listing_count(
    country_codes: list[str] | None = None,
    industry: str = "",
    partner_tiers: list[str] | None = None,
    language: str = "",
) -> int:
    url = build_list_url(country_codes, page=1, industry=industry,
                          partner_tiers=partner_tiers, language=language)
    html = fetch(url)
    if not html:
        return 0
    match = re.search(r"of\s+([\d,]+)(?:</[^>]+>)?\s*partner", html, re.I)
    if match:
        return int(match.group(1).replace(",", ""))
    return 0

# ---------------------------------------------------------------------------
# STEP 1: COLLECT PARTNER LINKS  (generator)
# ---------------------------------------------------------------------------

def collect_partner_links(
    country_codes: list[str] | None = None,
    max_pages: int = MAX_PAGES,
    limit: int | None = None,
    industry: str = "",
    partner_tiers: list[str] | None = None,
    language: str = "",
):
    partners: list[dict] = []
    seen_urls: set[str] = set()

    for page in range(1, max_pages + 1):
        if _stop_requested:
            break

        if limit and limit > 0 and len(partners) >= limit:
            break

        url = build_list_url(country_codes, page=page, industry=industry,
                              partner_tiers=partner_tiers, language=language)
        html = fetch(url)
        if not html:
            yield {"type": "listing_skip", "page": page}
            continue

        soup = BeautifulSoup(html, "html.parser")
        links = soup.find_all("a", href=re.compile(r"/partners/directory/partner/[^?]+$"))

        page_count = 0
        for a in links:
            if limit and limit > 0 and len(partners) >= limit:
                break
            href = a.get("href")
            full_url = href if href.startswith("http") else f"https://www.shopify.com{href}"
            if full_url in seen_urls:
                continue
            seen_urls.add(full_url)
            name_tag = a.find(["h3", "h2"])
            name = name_tag.get_text(strip=True) if name_tag else a.get_text(strip=True)[:80]
            partners.append({"name": name, "profile_url": full_url})
            page_count += 1

        yield {"type": "listing_page", "page": page, "found": page_count, "total": len(partners)}

        if page_count == 0 and page > 1:
            break

        polite_sleep()

    yield {"type": "collect_done", "total": len(partners), "partners": partners}

# ---------------------------------------------------------------------------
# STEP 2: SCRAPE PARTNER DETAIL
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"tel:([+0-9()\-.\s]{7,})")


def scrape_partner_detail(profile_url: str) -> dict:
    result = {
        "website": "",
        "email": "",
        "phone": "",
        "location": "",
    }

    html = fetch(profile_url)
    if not html:
        return result

    soup = BeautifulSoup(html, "html.parser")

    mailto = soup.find("a", href=re.compile(r"^mailto:"))
    if mailto:
        result["email"] = mailto["href"].replace("mailto:", "").strip()
    else:
        match = EMAIL_RE.search(html)
        if match:
            result["email"] = match.group(0)

    tel = soup.find("a", href=re.compile(r"^tel:"))
    if tel:
        result["phone"] = tel["href"].replace("tel:", "").strip()

    contact_links = soup.find_all("a", href=True)
    for a in contact_links:
        href = a["href"]
        if href.startswith("http") and "shopify.com" not in href:
            result["website"] = href
            break

    loc_label = soup.find(string=re.compile(r"Primary location", re.I))
    if loc_label:
        parent = loc_label.find_parent()
        if parent:
            sib = parent.find_next_sibling()
            if sib:
                result["location"] = sib.get_text(strip=True)

    if not result["location"]:
        loc_match = re.search(r'"location"\s*:\s*"([^"]+)"', html)
        if loc_match:
            result["location"] = loc_match.group(1)

    return result

# ---------------------------------------------------------------------------
# EXCEL I/O
# ---------------------------------------------------------------------------

COLUMNS = ["Name", "Website", "Email", "Phone", "Location", "Profile URL"]


def load_existing(path: str) -> tuple[Workbook, set]:
    if Path(path).exists():
        wb = load_workbook(path)
        ws = wb.active
        done_urls = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6 and row[5]:
                done_urls.add(row[5])
        return wb, done_urls

    wb = Workbook()
    ws = wb.active
    ws.title = "Partners"
    ws.append(COLUMNS)
    return wb, set()


def append_row(wb: Workbook, row: list):
    wb.active.append(row)


def save(wb: Workbook, path: str):
    wb.save(path)

# ---------------------------------------------------------------------------
# FULL SCRAPE GENERATOR
# ---------------------------------------------------------------------------

_completed_files: dict[str, str] = {}


def run_scrape(
    country_codes: list[str] | None = None,
    output_path: str = OUTPUT_FILE,
    limit: int | None = None,
    industry: str = "",
    partner_tiers: list[str] | None = None,
    language: str = "",
):
    yield {"type": "phase", "phase": "collecting",
           "message": "Collecting partner links from listing pages..."}

    partners: list[dict] = []
    for event in collect_partner_links(country_codes, limit=limit,
                                        industry=industry, partner_tiers=partner_tiers,
                                        language=language):
        if event["type"] == "collect_done":
            partners = event.get("partners", [])
        yield event

    if not partners:
        yield {"type": "error", "message": "No partners found for this country."}
        return

    if limit and limit > 0 and limit < len(partners):
        partners = partners[:limit]
        yield {"type": "limited", "total": len(partners)}

    yield {"type": "phase", "phase": "scraping",
           "message": f"Scraping details for {len(partners)} partners..."}

    wb, done_urls = load_existing(output_path)
    remaining = [p for p in partners if p["profile_url"] not in done_urls]
    skipped = len(partners) - len(remaining)

    yield {"type": "scrape_start", "total": len(remaining), "skipped": skipped}

    if not remaining:
        yield {"type": "already_done", "message": "All partners were already scraped."}
        save(wb, output_path)
        download_id = str(uuid.uuid4())
        _completed_files[download_id] = output_path
        yield {"type": "done", "download_id": download_id}
        return

    for i, partner in enumerate(remaining, 1):
        if _stop_requested:
            yield {"type": "stop", "current": i, "total": len(remaining)}
            break

        detail = scrape_partner_detail(partner["profile_url"])
        append_row(wb, [
            partner["name"],
            detail["website"],
            detail["email"],
            detail["phone"],
            detail["location"],
            partner["profile_url"],
        ])
        yield {"type": "scrape_progress", "current": i, "total": len(remaining),
               "name": partner["name"]}

        if i % 25 == 0:
            save(wb, output_path)
            yield {"type": "save", "current": i, "total": len(remaining)}

        polite_sleep()

    save(wb, output_path)

    download_id = str(uuid.uuid4())
    _completed_files[download_id] = output_path
    yield {"type": "done", "download_id": download_id}


def get_download_path(download_id: str) -> str | None:
    return _completed_files.get(download_id)


def cleanup_download(download_id: str):
    _completed_files.pop(download_id, None)

# ---------------------------------------------------------------------------
# CLI MAIN
# ---------------------------------------------------------------------------

def main():
    print("=== Step 1: Collecting partner links from listing pages ===")
    partners = []
    for event in collect_partner_links():
        if event["type"] == "collect_done":
            partners = event.get("partners", [])
        elif event["type"] == "listing_page":
            print(f"  Page {event['page']}: {event['found']} partners found (total: {event['total']})")

    print(f"\nCollected {len(partners)} partner links.\n")

    print("=== Step 2: Scraping partner detail pages ===")
    wb, done_urls = load_existing(OUTPUT_FILE)
    print(f"Already scraped: {len(done_urls)} (will skip these)")

    remaining = [p for p in partners if p["profile_url"] not in done_urls]
    print(f"Remaining to scrape: {len(remaining)}\n")

    for i, partner in enumerate(remaining, 1):
        if _stop_requested:
            print("Stopping early per user request.")
            break

        print(f"[{i}/{len(remaining)}] {partner['name']}")
        detail = scrape_partner_detail(partner["profile_url"])

        append_row(wb, [
            partner["name"],
            detail["website"],
            detail["email"],
            detail["phone"],
            detail["location"],
            partner["profile_url"],
        ])

        if i % 25 == 0:
            save(wb, OUTPUT_FILE)
            print(f"  ...progress saved ({i}/{len(remaining)})")

        polite_sleep()

    save(wb, OUTPUT_FILE)
    print(f"\nDone. Saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
